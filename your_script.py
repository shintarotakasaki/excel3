import streamlit as st
import fitz  # PyMuPDF
from openpyxl import load_workbook
import requests
import tempfile

# PDFからテキストを抽出する関数
def extract_text_from_pdf(pdf_file, rects):
    text_list = []
    labels = []
    
    # Streamlitでアップロードされたファイルを一時ファイルとして保存
    with fitz.open(stream=pdf_file.read(), filetype="pdf") as doc:  
        for rect_data in rects:
            x0, y0, x1, y1, label = rect_data
            rect = fitz.Rect(x0, y0, x1, y1)
            labels.append(label)
            text = ""
            for page in doc:
                extracted_text = page.get_text("text", clip=rect)
                text += extracted_text.strip() if extracted_text.strip() else ""
            text_list.append(text)
    return text_list, labels

# StreamlitによるGUI
def main():
    st.title("伝票作成アプリ")

    # PDFアップロード
    uploaded_pdf = st.file_uploader("PDFファイルをアップロードしてください", type=["pdf"])
    if uploaded_pdf:
        st.success(f"{uploaded_pdf.name} がアップロードされました")

        # 抽出範囲の指定
        rects = [
            (140, 175, 180, 190, 'AC9-1'),
            (140, 190, 180, 205, 'AC9'),
            (210, 190, 500, 205, 'AC11'),
            (140, 205, 180, 220, 'AC13'),
            (210, 205, 500, 220, 'AC15'),
            (100, 215, 140, 230, 'AC17'),
            (135, 220, 500, 230, 'AC19'),
            (105, 295, 250, 305, 'A11'),
            (400, 360, 500, 370, 'S11')
        ]

        # テキスト抽出
        text_list, labels = extract_text_from_pdf(uploaded_pdf, rects)
        st.write("抽出されたテキスト:")
        for label, text in zip(labels, text_list):
            st.write(f"**{label}**: {text}")

        # ユーザー入力
        syukka = st.text_input("出荷日を入力してください（例: 2024年11月18日）")
        buturyu = st.text_input("物流センターを入力してください")
        konpou = st.text_input("梱包数を入力してください")

        # Excel更新
        if st.button("Excelファイルを生成する"):
            # GitHubのリポジトリURL
            github_url = "https://github.com/yourusername/excel3/raw/main/伝票(規格品)_ラベル_指示書.xlsm"

            # ファイルをダウンロードして一時ファイルとして保存
            response = requests.get(github_url)
            if response.status_code == 200:
                with tempfile.NamedTemporaryFile(delete=False) as tmp_file:
                    tmp_file.write(response.content)
                    file_path = tmp_file.name
            
            try:
                wb = load_workbook(file_path, keep_vba=True)
                ws = wb.active
                
                # Excelファイルへの書き込み
                ws['AH3'] = syukka
                ws['AM9'] = buturyu
                ws['AB4'] = konpou + "梱包"

                zig_tok = ""
                for i, text in enumerate(text_list):
                    if labels[i] == 'AC9-1':
                        zig_tok = text
                    elif labels[i] == 'AC9':
                        ws[labels[i]] = zig_tok + '-' + text
                    elif labels[i] == 'AC11':
                        ws[labels[i]] = text + "様"
                    elif labels[i] == 'AC13':
                        ws[labels[i]] = '届け先：' + text
                    elif labels[i] == 'AC15':
                        ws[labels[i]] = text + "様" if text else "=AC11"
                    elif labels[i] == 'AC17':
                        ws[labels[i]] = '現場名：' + text
                    else:
                        ws[labels[i]] = text

                # 保存とダウンロード
                wb.save(file_path)
                st.success("Excelファイルが上書き保存されました！")                

                #ファイルをダウンロード
                st.write("保存されたファイルを以下のリンクからダウンロードしてください:")
                with open(file_path, "rb") as file:
                    st.download_button(
                        label="ダウンロードする",
                        data=file,
                        file_name="伝票(規格品)_ラベル_指示書.xlsm",
                        mime="application/vnd.ms-excel"
                    )
            except Exception as e:
                st.error(f"エラーが発生しました: {e}")                

if __name__ == "__main__":
    main()
